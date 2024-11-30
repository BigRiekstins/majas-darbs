import csv
from openpyxl import load_workbook


description_file = '/otto/python nedarbi/dip225-2-practical-task-BigRiekstins-main/description.xlsx'
data_file = '/otto/python nedarbi/dip225-2-practical-task-BigRiekstins-main/data.csv'


def load_region_codes(description_file):
    try:
        workbook = load_workbook(filename=description_file)
        sheet = workbook.active


        region_mapping = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 2 or not row[0] or not row[1]: 
                continue
            region_name, region_code = row[:2]
            region_mapping[str(region_name).strip()] = str(region_code).strip()
        print(f"Successfully loaded region mappings from {description_file}.")
        return region_mapping
    except FileNotFoundError:
        print(f"Error: The required file '{description_file}' was not found.")
        return {}
    except Exception as e:
        print(f"Error while processing the file '{description_file}': {e}")
        return {}


def sum_geo_count(data_file, region_code):
    try:
        with open(data_file, mode='r', newline='') as file:
            reader = csv.DictReader(file)
            total = 0

            for row in reader:
                if 'Region Code' in row and 'geo_count' in row:
                    if row['Region Code'].strip() == region_code:
                        try:
                            total += int(row['geo_count'])
                        except ValueError:
                            print(f"Invalid geo_count value in '{data_file}' for row: {row}")
                            continue
                else:
                    print(f"Error: The required columns 'Region Code' or 'geo_count' are missing in the file '{data_file}'.")
                    return 0
            print(f"Successfully calculated geo_count for region code {region_code} from '{data_file}'.")
            return total
    except FileNotFoundError:
        print(f"Error: The required file '{data_file}' was not found.")
        return 0
    except Exception as e:
        print(f"Error while processing the file '{data_file}': {e}")
        return 0


def main():
    
    region_mapping = load_region_codes(description_file)


    region_name = input("Enter region name: ").strip()


    region_code = region_mapping.get(region_name)


    if not region_code:
        print(0)
    else:
        total_geo_count = sum_geo_count(data_file, region_code)
        print(total_geo_count) 

if __name__ == "__main__":
    main()
